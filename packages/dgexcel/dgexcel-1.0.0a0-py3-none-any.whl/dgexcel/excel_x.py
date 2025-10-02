from string import printable
from typing import Union, Optional, List, Dict, Iterable
from datetime import datetime, timedelta, date
import os

from .base import ExcelBase
from .constants import BORDER

try:
    from openpyxl.utils.exceptions import IllegalCharacterError
    from openpyxl.styles import Border, Side, Alignment, Font
    import openpyxl
except ImportError:
    raise ImportError("openpyxl isn't installed. Run: pip install openpyxl")


class ExcelX(ExcelBase):
    def open_file(self, filename: str = None, rewrite=True):
        if filename and not filename.endswith('xlsx'):
            filename += '.xlsx'

        super().open_file(filename=filename)

        if os.path.exists(self.filename) and not rewrite:
            self.get_file()
        else:
            self.create_file()

        self.worksheets = {w.title: w for w in self.workbook.worksheets}
        self.current_worksheet = self.workbook.active

    def get_file(self):
        self.workbook = openpyxl.load_workbook(self.filename)

    def create_file(self):
        self.workbook = openpyxl.workbook.Workbook()

    def add_worksheet(self, worksheet: str):
        self.worksheets[worksheet] = self.workbook.create_sheet(worksheet)

    def rename_worksheet(self, worksheet_old: str, worksheet_new: str):
        ws = self.workbook[worksheet_old]
        ws.title = worksheet_new
        self.worksheets = {w.title: w for w in self.workbook.worksheets}
        self.select_worksheet(worksheet_new)

    def select_worksheet(self, worksheet):
        if not self.worksheets.get(worksheet):
            self.add_worksheet(worksheet)
        self.current_worksheet = self.worksheets.get(worksheet)

    def save_file(self, new_filename: str = None):
        filename = self.filename if not new_filename else new_filename
        self.workbook.save(filename)

    def add_cell_coords(self, coords: str, value: Union[None, str, float, int, date, datetime],
                        style: Optional[Dict] = None):
        if not self.current_worksheet:
            self.current_worksheet = self.workbook.active
        cell = self.current_worksheet[coords]
        self.add_cell(cell.row, cell.column, value, style)

    def add_cell(self, row: int, col: int, value: Union[None, str, float, int, date, datetime],
                 style: Optional[Dict] = None):
        if not self.current_worksheet:
            self.current_worksheet = self.workbook.active
        try:
            cell = self.current_worksheet.cell(row, col, value)
        except IllegalCharacterError:
            value = ''.join(c for c in value if c in printable)
            cell = self.current_worksheet.cell(row, col, value)
        if style:
            if style.get('border'):
                b = style.get('border')
                left = b.get('left', {})
                right = b.get('right', {})
                bottom = b.get('bottom', {})
                top = b.get('top', {})
                cell.border = Border(left=Side(style=left.get('style', None), color=left.get('color', 'FF000000')),
                                     right=Side(style=right.get('style', None), color=right.get('color', 'FF000000')),
                                     bottom=Side(style=bottom.get('style', None),
                                                 color=bottom.get('color', 'FF000000')),
                                     top=Side(style=top.get('style', None), color=top.get('color', 'FF000000')))
            if style.get('bold'):
                cell.font = Font(bold=True)
            if style.get('italic'):
                cell.font = Font(italic=True)
            if style.get('font'):
                f = style.get('font')
                cell.font = Font(name=f.get('font', 'Calibri'),
                                 size=f.get('size', 11),
                                 bold=f.get('bold', False),
                                 italic=f.get('italic', False),
                                 underline=f.get('underline', 'none'),
                                 strike=f.get('strike', False),
                                 color=f.get('color', '000000'))
            if style.get('alignment'):
                cell.alignment = Alignment(horizontal=style.get('alignment').get('horizontal', 'general'),
                                           vertical=style.get('alignment').get('vertical', 'bottom'),
                                           wrap_text=style.get('alignment').get('wrap_text', None))
            if style.get('number_format'):
                cell.number_format = style.get('number_format', 'general')

    def add_row(self, row: int, list_cells: Iterable[Union[None, str, float, int, date, datetime]],
                start_col: int = 1, styles: Union[None, List, Dict] = None):
        if isinstance(list_cells, dict):
            list_cells_ = list(list_cells.values())
        else:
            list_cells_ = list_cells
        for col, cell in enumerate(list_cells_):
            style = None
            if styles:
                if isinstance(styles, list):
                    style = styles[col]
                elif isinstance(styles, dict):
                    style = styles
            self.add_cell(row, start_col + col, cell, style)
        self.rows_qty += 1

    def add_column(self, col: int, list_cells: Iterable[Union[None, str, float, int, date, datetime]],
                   start_row: int = 1, styles: Union[None, List, Dict] = None):
        if isinstance(list_cells, dict):
            list_cells_ = list(list_cells.values())
        else:
            list_cells_ = list_cells
        for row, cell in enumerate(list_cells_):
            style = None
            if styles:
                if isinstance(styles, list):
                    style = styles[row]
                elif isinstance(styles, dict):
                    style = styles
            self.add_cell(start_row + row, col, cell, style)
            self.rows_qty += 1

    def add_list_rows(self, first_row: int, first_col: int, list_rows: Iterable[Union[Dict, List]],
                      styles: Union[None, List, Dict] = None):
        rows_cnt = len(list_rows)
        if styles:
            styles_cnt = len(styles)
        else:
            styles_cnt = 0
        for row_index, row_ in enumerate(list_rows):
            if isinstance(row_, dict):
                row = list(row_.values())
            else:
                row = row_
            if isinstance(styles, list) and styles_cnt == rows_cnt:
                self.add_row(first_row + row_index, row, first_col, styles[row_index])
            else:
                self.add_row(first_row + row_index, row, first_col, styles)

    def add_merge_cell(self, first_row: int, first_col: int, last_row: int, last_col: int,
                       value: Union[None, str, float, int, date, datetime], style: Optional[Dict] = None):
        self.add_cell(first_row, first_col, value, style)
        self.current_worksheet.merge_cells(start_row=first_row, start_column=first_col,
                                           end_row=last_row, end_column=last_col)

    def add_merge_cell_coords(self, first_coords: str, last_coords,
                              value: Union[None, str, float, int, date, datetime], style: Optional[Dict] = None):
        if not self.current_worksheet:
            self.current_worksheet = self.workbook.active
        first_cell = self.current_worksheet[first_coords]
        last_cell = self.current_worksheet[last_coords]
        self.add_merge_cell(first_cell.row, first_cell.column, last_cell.row, last_cell.column, value, style)

    def set_cell_dimension(self, row: int, col: int, height: Union[int, float] = 15, width: Union[int, float] = 10):
        cell = self.current_worksheet.cell(row, col)
        self.current_worksheet.column_dimensions[cell.column_letter].width = width
        self.current_worksheet.row_dimensions[row].height = height

    def set_rows_height(self, list_rows_height: Iterable[float], start_row: int = 1):
        for row, height in enumerate(list_rows_height):
            self.current_worksheet.row_dimensions[start_row + row].height = height

    def set_columns_width(self, list_columns_width: Iterable[float], start_col: int = 1):
        for col, width in enumerate(list_columns_width):
            cell = self.current_worksheet.cell(1, start_col + col)
            self.current_worksheet.column_dimensions[cell.column_letter].width = width

    def create_simple_excel(self, filename, data: Iterable[Union[Dict, List]], column_width: Optional[Iterable] = None,
                            column_names: Optional[Iterable] = None, styles: Union[None, List, Dict] = None,
                            worksheet: str = 'Sheet', rewrite: bool = True, start_row: int = 1):
        header_style = {'bold': True, 'border': BORDER.get('bottom'),
                        'alignment': {'horizontal': 'center', 'wrap_text': True}}

        self.open_file(filename, rewrite)
        if rewrite:
            self.rename_worksheet('Sheet', worksheet)
        self.select_worksheet(worksheet)

        if column_width:
            self.set_columns_width(column_width)

        _start_row = start_row

        if column_names:
            self.add_row(_start_row, column_names, styles=header_style)
            _start_row += 1

        if any(isinstance(e, dict) for e in data):
            self.add_list_rows(_start_row, 1, [list(x.values()) for x in data], styles)
        else:
            self.add_list_rows(_start_row, 1, data, styles)
        self.save_file()

    def get_max_row(self):
        if not self.current_worksheet:
            self.current_worksheet = self.workbook.active
        return self.current_worksheet.max_row

    def get_max_column(self):
        if not self.current_worksheet:
            self.current_worksheet = self.workbook.active
        return self.current_worksheet.max_column

    def read_file(self, excel_header: list, sheet_name: str = None,
               int_columns: list = None, date_columns: list = None, start_row: int = 1):
        excel_arr = []
        columns = excel_header

        ws = self.worksheets.get(sheet_name, self.current_worksheet)
        records = [[x.value for x in row] for row in ws.iter_rows(min_row=start_row)]

        int_columns = int_columns if int_columns else []
        date_columns = date_columns if date_columns else []

        for r in records:
            tmp_row = dict(zip(columns, r))

            for c in columns:
                if tmp_row.get(c) is not None:
                    if c in int_columns and not isinstance(tmp_row.get(c), int):
                        tmp_row[c] = int(tmp_row[c])
                    if c in date_columns and not isinstance(tmp_row.get(c), datetime):
                        try:
                            tmp_row[c] = self._parse_date(tmp_row[c])
                        except (TypeError, ValueError):
                            try:
                                tmp_row[c] = datetime(1899, 12, 31) + timedelta(days=tmp_row[c] - 1)
                            except (TypeError, ValueError):
                                tmp_row[c] = None
                    if c not in date_columns and c not in int_columns and not isinstance(tmp_row.get(c), str):
                        tmp_row[c] = str(tmp_row.get(c))
            excel_arr.append(tmp_row)

        return excel_arr