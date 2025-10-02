# Copyright CNRS/Inria/UCA
# Contributor(s): Eric Debreuve (since 2021)
#
# eric.debreuve@cnrs.fr
#
# This software is governed by the CeCILL  license under French law and
# abiding by the rules of distribution of free software.  You can  use,
# modify and/ or redistribute the software under the terms of the CeCILL
# license as circulated by CEA, CNRS and INRIA at the following URL
# "http://www.cecill.info".
#
# As a counterpart to the access to the source code and  rights to copy,
# modify and redistribute granted by the license, users are provided only
# with a limited warranty  and the software's author,  the holder of the
# economic rights,  and the successive licensors  have only  limited
# liability.
#
# In this respect, the user's attention is drawn to the risks associated
# with loading,  using,  modifying and/or developing or reproducing the
# software by the user in light of its specific status of free software,
# that may mean  that it is complicated to manipulate,  and  that  also
# therefore means  that it is reserved for developers  and  experienced
# professionals having in-depth computer knowledge. Users are therefore
# encouraged to load and test the software's suitability as regards their
# requirements in conditions enabling the security of their systems and/or
# data to be ensured and,  more generally, to use and operate it in the
# same conditions as regards security.
#
# The fact that you are presently reading this means that you have had
# knowledge of the CeCILL license and that you accept its terms.

"""
Quick and dirty work done for a one-shot use in a project.
"""

import dataclasses as dtcl
import datetime as dttm
import typing as h

import networkx as ntwx
from cell_tracking_BC.in_out.file.divisions import SaveCellDivisionsToXLSX
from cell_tracking_BC.in_out.file.track import SaveTrackDetailsToXLSX
from cell_tracking_BC.task.tracking.constant import TRACKING_AFFINITY_LABEL
from cell_tracking_BC.type.compartment.cell import state_e
from cell_tracking_BC.type.track.multiple.structured import tracks_t
from cell_tracking_BC.type.track.multiple.unstructured import unstructured_tracks_t
from cell_tracking_BC.type.track.single.unstructured import TIME_POINT
from openpyxl import load_workbook as WorkbookFromPath

# from cell_tracking_BC.standard.uid import Identity

INFINITY = 999999999


@dtcl.dataclass(repr=False, eq=False)
class fake_cell_t:
    uid: tuple[int, ...]
    centroid: tuple[int, ...]
    all_labels: tuple[int, ...]
    time_point: int
    label: h.Any
    state: state_e  # = state_e.living

    @classmethod
    def NewFromLabelsAndTimePoint(cls, labels, time_point):
        """"""
        all_labels = tuple(sorted(labels))
        return cls(
            uid=all_labels + (time_point,),
            centroid=all_labels + (time_point,),
            all_labels=all_labels,
            time_point=time_point,
            label=(time_point,) + all_labels,
            state=state_e.living,
        )

    # __repr__ = Identity
    def __repr__(self):
        """"""
        return str(self.uid)

    def __eq__(self, other):
        """"""
        return self.uid == other.uid

    def __hash__(self):
        """"""
        return hash(self.label)

    def __str__(self):
        """"""
        return (
            "cell("
            + str(self.all_labels)
            + " "
            + str(self.time_point)
            + " "
            + str(self.state)
            + ")"
        )


def RebuiltTracks(tracks_path: str, events_path: str, /) -> h.Any:
    """"""
    thread_tracks = SheetContent(tracks_path, sheet={"siblings": 0, "status": 0})
    divisions = SheetContent(events_path, sheet={"division times": None})

    DealWithIntegersInTrackStrings(thread_tracks)
    DealWithIntegersInDivisionStrings(divisions)

    invalid_tracks = []

    edges_per_label = {}
    unstructured_tracks = unstructured_tracks_t()
    already_processed = set()
    for label, (siblings, status) in enumerate(
        zip(thread_tracks["siblings"], thread_tracks["status"]), start=1
    ):
        if status != "Valid":
            invalid_tracks.append((label, siblings, status))
            continue
        if label in already_processed:
            continue

        if isinstance(siblings, int):
            labels = [label, siblings]
        elif isinstance(siblings, (list, tuple)):
            labels = [label] + list(siblings)
        else:
            labels = [label]
        already_processed |= set(labels)

        if isinstance(divisions[label - 1], str):
            assert labels.__len__() == 1
            source = ((labels[0],), 0)
            target = ((labels[0],), INFINITY)
            edges = ((source, target),)
        else:
            nodes = ForkingTrackNodes(labels, divisions)
            edges = ForkingTrackEdges(nodes)

        edges_per_label[label] = edges

        for edge in edges:
            source, target = edge
            AddEdgeToTrack(source, target, unstructured_tracks)

    valid_tracks = tracks_t.NewFromUnstructuredTracks(unstructured_tracks)

    dividing_cells = {}
    for valid_track in valid_tracks:
        MarkDividingCells(valid_track, dividing_cells)

    label_conversions = {}
    for valid_track in valid_tracks:
        for cells, label in valid_track.LabeledThreadIterator():
            # print("++++", label, "root", cells[0].uid, "leaf", cells[-1].uid)
            label_conversions[label] = cells[-1].all_labels[0]  # Leaf cell label.

    next_label_for_valid_s = max(label_conversions.keys()) + 1
    n_thread_tracks_total = thread_tracks["siblings"].__len__()

    assert sorted(label_conversions.keys()) == list(range(1, next_label_for_valid_s))
    assert sorted(label_conversions.values()) == list(
        set(range(1, n_thread_tracks_total + 1)).difference(
            _elm[0] for _elm in invalid_tracks
        )
    )

    return (
        thread_tracks,
        divisions,
        dividing_cells,
        valid_tracks,
        next_label_for_valid_s,
        label_conversions,
        invalid_tracks,
    )


def MarkDividingCells(valid_track, dividing_cells):
    """"""
    valid_track._dividing_marked = True
    if not isinstance(valid_track, ntwx.DiGraph):
        return

    for cells, label in valid_track.LabeledThreadIterator():
        for cell in cells:
            if cell.state == state_e.discarded:
                continue

            if valid_track.CellSuccessors(cell).__len__() > 1:
                if label not in dividing_cells:
                    dividing_cells[label] = []
                if cell.uid not in dividing_cells[label]:
                    dividing_cells[label].append(cell.uid)
                cell.state = state_e.dividing
            else:
                if (label in dividing_cells) and (cell.uid in dividing_cells[label]):
                    del dividing_cells[label][cell.uid]
                cell.state = state_e.living


def SaveAndReloadTracks(analysis, label_conversions, folder, suffix):
    """"""
    path = folder / f"tracks-{suffix}.xlsx"
    SaveTrackDetailsToXLSX(path, analysis, can_overwrite=True)

    new_thread_tracks = SheetContent(str(path), sheet={"siblings": 0, "status": 0})
    DealWithIntegersInTrackStrings(new_thread_tracks)
    CorrectTrackLabels(new_thread_tracks, label_conversions)

    return new_thread_tracks


def SaveAndReloadDivisions(
    analysis, label_conversions, dividing_cells, unwanted, overlapping, folder, suffix
):
    """"""
    new_dividing_cells = dividing_cells  # {label_conversions[_key]: _vle for _key, _vle in dividing_cells.items()}

    path = folder / f"cell_events-{suffix}.xlsx"
    SaveCellDivisionsToXLSX(
        path, analysis, new_dividing_cells, unwanted, overlapping, label_conversions
    )

    new_thread_tracks = SheetContent(str(path), sheet={"division times": None})
    DealWithIntegersInDivisionStrings(new_thread_tracks)
    # CorrectTrackLabels(new_thread_tracks, label_conversions)

    return new_thread_tracks


def SheetContent(
    path: str,
    /,
    *,
    sheet: (
        str | h.Sequence[str] | dict[str, int | h.Sequence[int] | None] | None
    ) = None,
) -> h.Any:
    """"""
    output = {}

    if isinstance(sheet, str):
        sheet = {sheet: None}
    elif isinstance(sheet, h.Sequence):
        sheet = {_elm: None for _elm in sheet}
    elif isinstance(sheet, dict):
        standard = {}
        for name, col_idx in sheet.items():
            if col_idx is None:
                standard[name] = None
            elif isinstance(col_idx, int):
                standard[name] = (col_idx,)
            else:
                standard[name] = col_idx
        sheet = standard
    col_idx = None

    workbook = WorkbookFromPath(filename=path)
    for current_name in workbook.sheetnames:
        if (sheet is None) or (current_name in sheet):
            current_sheet = workbook[current_name]
            if isinstance(sheet, dict):
                col_idx = sheet[current_name]

            content = []
            for row in current_sheet.iter_rows(values_only=True):
                if (row[0] == "END") and all(_elm is None for _elm in row[1:]):
                    break

                selected = [
                    _elm
                    for _idx, _elm in enumerate(row)
                    if (col_idx is None) or (_idx in col_idx)
                ]

                if None in selected:
                    first_none = selected.index(None)
                    del selected[first_none:]
                    if selected.__len__() == 0:
                        selected = [None]

                if (col_idx is not None) and (col_idx.__len__() == 1):
                    selected = selected[0]
                content.append(selected)

            output[current_name] = content

    if output.__len__() > 1:
        return output

    if output.__len__() > 0:
        return output[tuple(output.keys())[0]]

    raise ValueError(
        f"{path}: No sheet(s) with name(s): {str(sorted(sheet.keys()))[1:-1]}; "
        f"Expected={workbook.sheetnames}."
    )


def DealWithIntegersInTrackStrings(thread_tracks):
    """"""
    for idx, cell in enumerate(thread_tracks["siblings"]):
        if isinstance(cell, str):
            values = cell.split(", ")
            converted = []
            for value in values:
                try:
                    value = int(value)
                except ValueError:
                    pass
                converted.append(value)
            if converted.__len__() == 1:
                converted = converted[0]
            thread_tracks["siblings"][idx] = converted


def DealWithIntegersInDivisionStrings(divisions):
    """"""
    for idx, values in enumerate(divisions):
        if isinstance(values, int):
            divisions[idx] = [values]
        elif isinstance(values, str):
            pass
        elif isinstance(values[0], str):
            assert values.__len__() == 1, values
            divisions[idx] = values[0]
        else:
            converted = []
            for value in values:
                if isinstance(value, str):
                    try:
                        value = int(value)
                    except ValueError:
                        pass
                converted.append(value)
            divisions[idx] = converted


def AddNodeToTrackIfNeeded(source, unstructured):
    """"""
    if isinstance(source, fake_cell_t):
        return source

    all_labels = tuple(source[0])
    time_point = source[1]

    source = fake_cell_t.NewFromLabelsAndTimePoint(all_labels, time_point)

    should_add_source = True
    for node in unstructured:
        if node.uid == source.uid:
            source = node
            should_add_source = False
            break

    if should_add_source:
        time_point_src = {TIME_POINT: time_point}
        unstructured.add_node(source, **time_point_src)

    return source


def AddEdgeToTrack(source, target, unstructured):
    """"""
    source = AddNodeToTrackIfNeeded(source, unstructured)
    target = AddNodeToTrackIfNeeded(target, unstructured)

    affinity = {TRACKING_AFFINITY_LABEL: 1.0}
    unstructured.add_edge(source, target, **affinity)


def ForkingTrackNodes(labels, divisions):
    """"""
    output = [(tuple(sorted(set(labels))), 0)] + [
        ((_elm,), INFINITY) for _elm in labels
    ]

    for sibling in labels:
        others = set(labels).difference([sibling])
        for time in divisions[sibling - 1]:
            tangent_s = TangentLabelsOfDivision(
                sibling, time, others, divisions, labels
            )
            node = (tuple(sorted({sibling} | set(tangent_s))), time)
            if node not in output:
                output.append(node)

    output = sorted(output, key=lambda _elm: _elm[1])

    return output


def ForkingTrackEdges(nodes):
    """"""
    output = []

    while nodes.__len__() > 1:
        node = nodes.pop(nodes.__len__() - 1)
        neighbors = MinimalSubset(node[0], nodes)
        for neighbor in neighbors:
            assert neighbor != node
            if node[1] > neighbor[1]:
                output.append((neighbor, node))
            else:
                output.append((node, neighbor))

    return output


def TangentLabelsOfDivision(
    sibling: int, time: int, siblings, divisions, labels, /
) -> list[int]:
    """"""
    output = []

    for current in siblings:
        if time in divisions[current - 1]:
            output.append(current)

    if output.__len__() == 0:
        raise RuntimeError(
            f"No siblings for division: {sibling} {time} {siblings} {labels}"
        )

    return output


def MinimalSubset(labels, nodes):
    """"""
    output = []

    for node in nodes:
        if set(node[0]).issuperset(labels):
            output.append(node)

    if output.__len__() == 0:
        raise RuntimeError(f"No minimal subsets: {labels} {nodes}")

    min_length = min(_elm[0].__len__() for _elm in output)
    output = [_elm for _elm in output if _elm[0].__len__() == min_length]

    if output.__len__() == 0:
        raise RuntimeError(f"No minimal subsets: {labels} {nodes}")

    max_time = max(_elm[1] for _elm in output)
    output = [_elm for _elm in output if _elm[1] == max_time]

    if output.__len__() == 0:
        raise RuntimeError(f"No minimal subsets: {labels} {nodes}")

    return output


def CorrectTrackLabels(thread_tracks, label_conversions):
    """"""
    thread_tracks_corrected = thread_tracks["siblings"].__len__() * [None]

    for idx, labels in enumerate(thread_tracks["siblings"]):
        if labels is None:
            thread_tracks_corrected[idx] = "Not in GT"
            # thread_tracks["siblings"][idx] = "Not in GT"
            continue

        if isinstance(labels, str):
            thread_tracks_corrected[idx] = labels
            continue

        if isinstance(labels, int):
            labels = (labels,)
        corrected = []
        for label in labels:
            corrected.append(label_conversions[label])
        if corrected.__len__() == 1:
            corrected = corrected[0]
        thread_tracks_corrected[idx] = corrected
        # thread_tracks["siblings"][idx] = corrected

    thread_tracks["siblings"] = thread_tracks_corrected
