import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
import json
import os

from . import HeadType
from .TableException import (
    FileNotFoundError as TableFileNotFoundError,
    InvalidFileFormatError,
    EmptyFileError,
    MissingHeaderMarkerError,
    MissingDataMarkerError
)


class Head:

    def __init__(self, sheet: Worksheet, row: tuple[Cell, ...], filename: str = None) -> None:
        self.sheet = sheet
        self.column = len(row)
        self.filename = filename
        self.head = HeadType.headCreater(row[0], filename)
        self.headList: list[HeadType.HeadType] = [self.head] * self.column

    def getCellMaxCol(self, cell: Cell) -> int:
        for i in self.sheet.merged_cells.ranges:
            if cell.coordinate in i:
                return i.max_col - 1
        return cell.column - 1

    def rowParser(self, row: tuple[Cell, ...]) -> None:
        i = 0
        while i < self.column:
            if row[i].value:
                h = HeadType.headCreater(row[i], self.filename)
                j = self.getCellMaxCol(row[i])
                self.headList[i].addChild(h)
                self.headList[i:j] = [h] * (j - i)
                i = j
            else:
                i += 1


def getData(filename: str) -> HeadType.data:
    # Validate input
    if not filename:
        raise ValueError("Filename cannot be empty")

    if not os.path.exists(filename):
        raise TableFileNotFoundError(filename)

    if not filename.lower().endswith(('.xlsx', '.xlsm')):
        raise InvalidFileFormatError(filename)

    # Load workbook
    try:
        workbook = openpyxl.load_workbook(filename)
    except Exception as e:
        raise InvalidFileFormatError(filename) from e

    if not workbook.active:
        raise EmptyFileError(filename)

    iter_rows = workbook.active.iter_rows()

    # Check first row
    try:
        first_row = next(iter_rows)
    except StopIteration:
        raise EmptyFileError(filename)

    # Validate #head marker
    first_cell = first_row[0]
    if first_cell.value != "#head":
        raise MissingHeaderMarkerError(
            first_cell,
            str(first_cell.value) if first_cell.value else "empty",
            filename
        )

    # Parse header
    head = Head(workbook.active, first_row[1:], filename)

    # Parse rows
    isData = False
    dataRoot = None

    for row in iter_rows:
        if row[0].value == "#":
            continue
        elif row[0].value == "#data":
            isData = True
            dataRoot = head.head.parsetData(row[1:], True, filename)
            continue

        if isData:
            head.head.parsetData(row[1:], False, filename)
        else:
            head.rowParser(row[1:])

    # Validate data was found
    if dataRoot is None:
        raise MissingDataMarkerError(filename)

    return dataRoot


def getJson(filename: str, indent: int = 2) -> str:
    """Convert Excel file to JSON string

    Args:
        filename: Path to Excel file
        indent: JSON indentation level (default 2)

    Returns:
        Formatted JSON string
    """
    return json.dumps(getData(filename), indent=indent)


if __name__ == "__main__":
    print(getJson("Table.xlsx"))
