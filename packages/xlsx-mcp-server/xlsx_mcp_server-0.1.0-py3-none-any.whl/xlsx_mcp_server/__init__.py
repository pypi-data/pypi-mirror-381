#!/usr/bin/env python3
"""
Excel MCP Server based on FastMCP
提供完整的 Excel 文件操作功能
"""

from fastmcp import FastMCP
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any, Optional
import tempfile
import os

# 创建 MCP 应用
mcp = FastMCP("Excel Server")

# 内存中的工作簿缓存
workbook_cache = {}

def get_workbook(workbook_id: str) -> Optional[Workbook]:
    """从缓存获取工作簿"""
    return workbook_cache.get(workbook_id)

def save_workbook(workbook_id: str, workbook: Workbook):
    """保存工作簿到缓存"""
    workbook_cache[workbook_id] = workbook

@mcp.tool()
def open_workbook(file_path: str, workbook_id: Optional[str] = None) -> Dict[str, Any]:
    """打开或创建 Excel 工作簿
    
    Args:
        file_path: Excel 文件路径
        workbook_id: 工作簿ID，默认为文件路径
        
    Returns:
        工作簿信息和工作表列表
    """
    try:
        wb_id = workbook_id or file_path
        
        if os.path.exists(file_path):
            workbook = openpyxl.load_workbook(file_path)
        else:
            workbook = Workbook()
            # 删除默认创建的空工作表（如果有多个）
            while len(workbook.sheetnames) > 1:
                workbook.remove(workbook.worksheets[-1])
        
        save_workbook(wb_id, workbook)
        
        return {
            "status": "success",
            "workbook_id": wb_id,
            "file_path": file_path,
            "sheets": workbook.sheetnames,
            "message": "Workbook opened successfully" if os.path.exists(file_path) else "New workbook created"
        }
    except Exception as e:
        return {"status": "error", "message": f"Failed to open workbook: {str(e)}"}

@mcp.tool()
def create_workbook(workbook_id: str, sheets: Optional[List[str]] = None) -> Dict[str, Any]:
    """创建新的工作簿
    
    Args:
        workbook_id: 工作簿ID
        sheets: 要创建的工作表名称列表
        
    Returns:
        工作簿信息
    """
    try:
        workbook = Workbook()
        
        # 重命名默认工作表
        if sheets and len(sheets) > 0:
            workbook.active.title = sheets[0]
            # 创建额外的工作表
            for sheet_name in sheets[1:]:
                workbook.create_sheet(sheet_name)
        else:
            # 使用默认工作表名称
            sheets = workbook.sheetnames
        
        save_workbook(workbook_id, workbook)
        
        return {
            "status": "success",
            "workbook_id": workbook_id,
            "sheets": sheets,
            "message": "New workbook created successfully"
        }
    except Exception as e:
        return {"status": "error", "message": f"Failed to create workbook: {str(e)}"}

@mcp.tool()
def list_sheets(workbook_id: str) -> Dict[str, Any]:
    """列出工作簿中的所有工作表
    
    Args:
        workbook_id: 工作簿ID
        
    Returns:
        工作表列表
    """
    workbook = get_workbook(workbook_id)
    if not workbook:
        return {"status": "error", "message": f"Workbook {workbook_id} not found"}
    
    return {
        "status": "success",
        "workbook_id": workbook_id,
        "sheets": workbook.sheetnames
    }

@mcp.tool()
def read_cell(workbook_id: str, sheet_name: str, cell: str) -> Dict[str, Any]:
    """读取指定单元格的值
    
    Args:
        workbook_id: 工作簿ID
        sheet_name: 工作表名称
        cell: 单元格引用（如 "A1"）
        
    Returns:
        单元格信息
    """
    workbook = get_workbook(workbook_id)
    if not workbook:
        return {"status": "error", "message": f"Workbook {workbook_id} not found"}
    
    try:
        sheet = workbook[sheet_name]
        cell_obj = sheet[cell]
        
        return {
            "status": "success",
            "workbook_id": workbook_id,
            "sheet": sheet_name,
            "cell": cell,
            "value": cell_obj.value,
            "has_formula": cell_obj.data_type == 'f',
            "formula": cell_obj.value if cell_obj.data_type == 'f' else None
        }
    except Exception as e:
        return {"status": "error", "message": f"Failed to read cell: {str(e)}"}

@mcp.tool()
def read_range(workbook_id: str, sheet_name: str, range_ref: str) -> Dict[str, Any]:
    """读取指定范围的数据
    
    Args:
        workbook_id: 工作簿ID
        sheet_name: 工作表名称
        range_ref: 范围引用（如 "A1:B10"）
        
    Returns:
        范围数据
    """
    workbook = get_workbook(workbook_id)
    if not workbook:
        return {"status": "error", "message": f"Workbook {workbook_id} not found"}
    
    try:
        sheet = workbook[sheet_name]
        data = []
        
        for row in sheet[range_ref]:
            row_data = []
            for cell in row:
                row_data.append({
                    "value": cell.value,
                    "cell": cell.coordinate,
                    "has_formula": cell.data_type == 'f'
                })
            data.append(row_data)
        
        return {
            "status": "success",
            "workbook_id": workbook_id,
            "sheet": sheet_name,
            "range": range_ref,
            "data": data
        }
    except Exception as e:
        return {"status": "error", "message": f"Failed to read range: {str(e)}"}

@mcp.tool()
def write_cell(workbook_id: str, sheet_name: str, cell: str, value: Any) -> Dict[str, Any]:
    """写入单元格值
    
    Args:
        workbook_id: 工作簿ID
        sheet_name: 工作表名称
        cell: 单元格引用
        value: 要写入的值
        
    Returns:
        操作结果
    """
    workbook = get_workbook(workbook_id)
    if not workbook:
        return {"status": "error", "message": f"Workbook {workbook_id} not found"}
    
    try:
        sheet = workbook[sheet_name]
        sheet[cell] = value
        
        return {
            "status": "success",
            "workbook_id": workbook_id,
            "sheet": sheet_name,
            "cell": cell,
            "value_set": value,
            "message": "Cell value set successfully"
        }
    except Exception as e:
        return {"status": "error", "message": f"Failed to write cell: {str(e)}"}

@mcp.tool()
def set_formula(workbook_id: str, sheet_name: str, cell: str, formula: str) -> Dict[str, Any]:
    """设置单元格公式
    
    Args:
        workbook_id: 工作簿ID
        sheet_name: 工作表名称
        cell: 单元格引用
        formula: Excel公式（以=开头）
        
    Returns:
        操作结果
    """
    workbook = get_workbook(workbook_id)
    if not workbook:
        return {"status": "error", "message": f"Workbook {workbook_id} not found"}
    
    try:
        sheet = workbook[sheet_name]
        sheet[cell] = formula
        
        return {
            "status": "success",
            "workbook_id": workbook_id,
            "sheet": sheet_name,
            "cell": cell,
            "formula_set": formula,
            "message": "Formula set successfully"
        }
    except Exception as e:
        return {"status": "error", "message": f"Failed to set formula: {str(e)}"}

@mcp.tool()
def write_dataframe(workbook_id: str, sheet_name: str, data: List[Dict], start_cell: str = "A1") -> Dict[str, Any]:
    """将数据框写入工作表
    
    Args:
        workbook_id: 工作簿ID
        sheet_name: 工作表名称
        data: 数据列表（字典列表）
        start_cell: 起始单元格
        
    Returns:
        操作结果
    """
    workbook = get_workbook(workbook_id)
    if not workbook:
        return {"status": "error", "message": f"Workbook {workbook_id} not found"}
    
    try:
        # 确保工作表存在
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)
        
        sheet = workbook[sheet_name]
        
        # 写入表头
        if data:
            headers = list(data[0].keys())
            for col_idx, header in enumerate(headers):
                col_letter = get_column_letter(col_idx + 1)
                cell_ref = f"{col_letter}{start_cell[1:] if start_cell[1:].isdigit() else 1}"
                sheet[cell_ref] = header
            
            # 写入数据
            for row_idx, row_data in enumerate(data):
                for col_idx, key in enumerate(headers):
                    col_letter = get_column_letter(col_idx + 1)
                    row_num = row_idx + 2  # +2 因为第一行是表头，Excel从1开始
                    cell_ref = f"{col_letter}{row_num}"
                    sheet[cell_ref] = row_data.get(key)
        
        return {
            "status": "success",
            "workbook_id": workbook_id,
            "sheet": sheet_name,
            "rows_written": len(data),
            "columns": len(data[0].keys()) if data else 0,
            "message": "Data written successfully"
        }
    except Exception as e:
        return {"status": "error", "message": f"Failed to write dataframe: {str(e)}"}

@mcp.tool()
def create_sheet(workbook_id: str, sheet_name: str) -> Dict[str, Any]:
    """创建新工作表
    
    Args:
        workbook_id: 工作簿ID
        sheet_name: 新工作表名称
        
    Returns:
        操作结果
    """
    workbook = get_workbook(workbook_id)
    if not workbook:
        return {"status": "error", "message": f"Workbook {workbook_id} not found"}
    
    try:
        workbook.create_sheet(sheet_name)
        return {
            "status": "success",
            "workbook_id": workbook_id,
            "sheet_created": sheet_name,
            "all_sheets": workbook.sheetnames,
            "message": f"Sheet '{sheet_name}' created successfully"
        }
    except Exception as e:
        return {"status": "error", "message": f"Failed to create sheet: {str(e)}"}

@mcp.tool()
def delete_sheet(workbook_id: str, sheet_name: str) -> Dict[str, Any]:
    """删除工作表
    
    Args:
        workbook_id: 工作簿ID
        sheet_name: 要删除的工作表名称
        
    Returns:
        操作结果
    """
    workbook = get_workbook(workbook_id)
    if not workbook:
        return {"status": "error", "message": f"Workbook {workbook_id} not found"}
    
    try:
        if sheet_name not in workbook.sheetnames:
            return {"status": "error", "message": f"Sheet '{sheet_name}' not found"}
        
        if len(workbook.sheetnames) <= 1:
            return {"status": "error", "message": "Cannot delete the last sheet in workbook"}
        
        sheet = workbook[sheet_name]
        workbook.remove(sheet)
        
        return {
            "status": "success",
            "workbook_id": workbook_id,
            "sheet_deleted": sheet_name,
            "remaining_sheets": workbook.sheetnames,
            "message": f"Sheet '{sheet_name}' deleted successfully"
        }
    except Exception as e:
        return {"status": "error", "message": f"Failed to delete sheet: {str(e)}"}

@mcp.tool()
def save_workbook_file(workbook_id: str, file_path: Optional[str] = None) -> Dict[str, Any]:
    """保存工作簿到文件
    
    Args:
        workbook_id: 工作簿ID
        file_path: 文件保存路径（可选）
        
    Returns:
        操作结果
    """
    workbook = get_workbook(workbook_id)
    if not workbook:
        return {"status": "error", "message": f"Workbook {workbook_id} not found"}
    
    try:
        save_path = file_path or workbook_id
        workbook.save(save_path)
        
        return {
            "status": "success",
            "workbook_id": workbook_id,
            "file_path": save_path,
            "message": f"Workbook saved to {save_path}"
        }
    except Exception as e:
        return {"status": "error", "message": f"Failed to save workbook: {str(e)}"}

@mcp.tool()
def close_workbook(workbook_id: str) -> Dict[str, Any]:
    """关闭工作簿并从缓存中移除
    
    Args:
        workbook_id: 工作簿ID
        
    Returns:
        操作结果
    """
    if workbook_id in workbook_cache:
        del workbook_cache[workbook_id]
        return {
            "status": "success",
            "message": f"Workbook {workbook_id} closed and removed from cache"
        }
    else:
        return {"status": "error", "message": f"Workbook {workbook_id} not found in cache"}

@mcp.tool()
def list_workbooks() -> Dict[str, Any]:
    """列出当前缓存中的所有工作簿
    
    Returns:
        工作簿列表
    """
    return {
        "status": "success",
        "workbooks": list(workbook_cache.keys()),
        "count": len(workbook_cache)
    }

@mcp.tool()
def get_sheet_info(workbook_id: str, sheet_name: str) -> Dict[str, Any]:
    """获取工作表的详细信息
    
    Args:
        workbook_id: 工作簿ID
        sheet_name: 工作表名称
        
    Returns:
        工作表信息
    """
    workbook = get_workbook(workbook_id)
    if not workbook:
        return {"status": "error", "message": f"Workbook {workbook_id} not found"}
    
    try:
        sheet = workbook[sheet_name]
        max_row = sheet.max_row
        max_column = sheet.max_column
        
        return {
            "status": "success",
            "workbook_id": workbook_id,
            "sheet_name": sheet_name,
            "dimensions": {
                "max_row": max_row,
                "max_column": max_column,
                "data_range": f"A1:{get_column_letter(max_column)}{max_row}" if max_row > 0 else "Empty"
            },
            "is_empty": max_row == 0
        }
    except Exception as e:
        return {"status": "error", "message": f"Failed to get sheet info: {str(e)}"}

# 主函数，用于作为入口点
def main():
    # 运行 MCP 服务器
    mcp.run(transport="stdio")

if __name__ == "__main__":
    main()