C=None
from._logger import logger
from sixgill.definitions import SystemVariables,ProfileVariables,Parameters,Constants
import pandas as a,os
from openpyxl import load_workbook
def perform_pt_analysis(self,study_name=C,thp=C,q_gas=C,q_oil=C,q_water=C,api=C,gg=C,gl_depth=C,gl_rate=C):
	e=False;d='openpyxl';c='wellfluid';b=gl_depth;Z='GLV';V=gl_rate;U=study_name;Q=gg;P=api;O=q_water;L=q_gas;K=q_oil;J=thp;I='LiquidLoadingGasRate';H='VolumeFlowrateGasStockTank';G='Pressure';F=True;A=self
	try:
		if b is not C:A.model.set_value(context=Z,parameter=Parameters.GasLiftInjection.TOPMEASUREDDEPTH,value=b)
		if V is not C:
			if V<=0:A.model.set_value(context=Z,parameter=Parameters.GasLiftInjection.GASRATE,value=1/1000000)
			else:A.model.set_value(context=Z,parameter=Parameters.GasLiftInjection.GASRATE,value=V/1000000)
		f=[L,K,O,P,Q];W=J is C;X=all(A is C for A in f);g=L is C;h=K is C
		if P is C:P=A.api
		if Q is C:Q=A.gg
		if X:
			D={Parameters.PTProfileSimulation.OUTLETPRESSURE:J,Parameters.PTProfileSimulation.GASFLOWRATE:A.q_gas/1000000,Parameters.PTProfileSimulation.FLOWRATETYPE:Constants.FlowRateType.GASFLOWRATE,Parameters.PTProfileSimulation.CALCULATEDVARIABLE:Constants.CalculatedVariable.INLETPRESSURE};E=[ProfileVariables.TEMPERATURE,ProfileVariables.PRESSURE,ProfileVariables.ELEVATION,ProfileVariables.TOTAL_DISTANCE,ProfileVariables.VOLUME_FLOWRATE_GAS_STOCKTANK,ProfileVariables.VOLUME_FLOWRATE_OIL_STOCKTANK,ProfileVariables.VOLUME_FLOWRATE_WATER_STOCKTANK]
			if A.gas_well==F:E.append(ProfileVariables.LIQUID_LOADING_GAS_RATE)
			R,B=A._pt_analysis_for_vlp(parameters=D,profile_variables=E);B[G]=(B[G]-1.01325)/.980665;B[H]=B[H]*1000000
			if A.gas_well==F:B[I]=B[I]*1000000
		elif W:
			M=L/K;N=O/(O+K)*100;A.model.set_values(context=c,dict={Parameters.BlackOilFluid.GOR:M,Parameters.BlackOilFluid.WATERCUT:N,Parameters.BlackOilFluid.API:P,Parameters.BlackOilFluid.GASSPECIFICGRAVITY:Q});E=[ProfileVariables.TEMPERATURE,ProfileVariables.PRESSURE,ProfileVariables.ELEVATION,ProfileVariables.TOTAL_DISTANCE,ProfileVariables.VOLUME_FLOWRATE_GAS_STOCKTANK,ProfileVariables.VOLUME_FLOWRATE_OIL_STOCKTANK,ProfileVariables.VOLUME_FLOWRATE_WATER_STOCKTANK]
			if A.gas_well==F:E.append(ProfileVariables.LIQUID_LOADING_GAS_RATE)
			D=D={Parameters.PTProfileSimulation.INLETPRESSURE:A.reservoir_pressure,Parameters.PTProfileSimulation.GASFLOWRATE:A.q_gas/1000000,Parameters.PTProfileSimulation.FLOWRATETYPE:Constants.FlowRateType.GASFLOWRATE,Parameters.PTProfileSimulation.CALCULATEDVARIABLE:Constants.CalculatedVariable.OUTLETPRESSURE};R,B=A._pt_analysis_for_vlp(parameters=D,profile_variables=E);B[G]=(B[G]-1.01325)/.980665;B[H]=B[H]*1000000
			if A.gas_well==F:B[I]=B[I]*1000000
		elif W and X:R,B=A._pt_analysis_for_vlp()
		elif not W and not X:
			try:N=O/(O+K)*100
			except:N=A.wc
			try:M=L/K
			except:M=A.gor
			A.model.set_values(context=c,dict={Parameters.BlackOilFluid.GOR:M,Parameters.BlackOilFluid.WATERCUT:N,Parameters.BlackOilFluid.API:P,Parameters.BlackOilFluid.GASSPECIFICGRAVITY:Q});E=[ProfileVariables.TEMPERATURE,ProfileVariables.PRESSURE,ProfileVariables.ELEVATION,ProfileVariables.TOTAL_DISTANCE,ProfileVariables.VOLUME_FLOWRATE_GAS_STOCKTANK,ProfileVariables.VOLUME_FLOWRATE_OIL_STOCKTANK,ProfileVariables.VOLUME_FLOWRATE_WATER_STOCKTANK]
			if A.gas_well==F:E.append(ProfileVariables.LIQUID_LOADING_GAS_RATE)
			try:D=D={Parameters.PTProfileSimulation.OUTLETPRESSURE:J,Parameters.PTProfileSimulation.GASFLOWRATE:L/1000000,Parameters.PTProfileSimulation.FLOWRATETYPE:Constants.FlowRateType.GASFLOWRATE,Parameters.PTProfileSimulation.CALCULATEDVARIABLE:Constants.CalculatedVariable.INLETPRESSURE}
			except:D=D={Parameters.PTProfileSimulation.OUTLETPRESSURE:J,Parameters.PTProfileSimulation.GASFLOWRATE:A.q_gas/1000000,Parameters.PTProfileSimulation.FLOWRATETYPE:Constants.FlowRateType.GASFLOWRATE,Parameters.PTProfileSimulation.CALCULATEDVARIABLE:Constants.CalculatedVariable.INLETPRESSURE}
			R,B=A._pt_analysis_for_vlp(parameters=D,profile_variables=E);B[G]=(B[G]-1.01325)/.980665;B[H]=B[H]*1000000
			if A.gas_well==F:B[I]=B[I]*1000000
		elif not g:
			if J is C:J=A.thp
			M=A.gor;N=A.wc;E=[ProfileVariables.TEMPERATURE,ProfileVariables.PRESSURE,ProfileVariables.ELEVATION,ProfileVariables.TOTAL_DISTANCE,ProfileVariables.VOLUME_FLOWRATE_GAS_STOCKTANK,ProfileVariables.VOLUME_FLOWRATE_OIL_STOCKTANK,ProfileVariables.VOLUME_FLOWRATE_WATER_STOCKTANK]
			if A.gas_well==F:E.append(ProfileVariables.LIQUID_LOADING_GAS_RATE)
			D=D={Parameters.PTProfileSimulation.INLETPRESSURE:A.reservoir_pressure,Parameters.PTProfileSimulation.GASFLOWRATE:L/1000000,Parameters.PTProfileSimulation.FLOWRATETYPE:Constants.FlowRateType.GASFLOWRATE,Parameters.PTProfileSimulation.CALCULATEDVARIABLE:Constants.CalculatedVariable.OUTLETPRESSURE};R,B=A._pt_analysis_for_vlp(parameters=D,profile_variables=E);B[G]=(B[G]-1.01325)/.980665;B[H]=B[H]*1000000
			if A.gas_well==F:B[I]=B[I]*1000000
		elif not h:
			if J is C:J=A.thp
			M=A.gor;N=A.wc;i=K/A.q_oil*(A.q_oil+A.q_water);E=[ProfileVariables.TEMPERATURE,ProfileVariables.PRESSURE,ProfileVariables.ELEVATION,ProfileVariables.TOTAL_DISTANCE,ProfileVariables.VOLUME_FLOWRATE_GAS_STOCKTANK,ProfileVariables.VOLUME_FLOWRATE_OIL_STOCKTANK,ProfileVariables.VOLUME_FLOWRATE_WATER_STOCKTANK]
			if A.gas_well==F:E.append(ProfileVariables.LIQUID_LOADING_GAS_RATE)
			D=D={Parameters.PTProfileSimulation.INLETPRESSURE:A.reservoir_pressure,Parameters.PTProfileSimulation.LIQUIDFLOWRATE:i,Parameters.PTProfileSimulation.FLOWRATETYPE:Constants.FlowRateType.LIQUIDFLOWRATE,Parameters.PTProfileSimulation.CALCULATEDVARIABLE:Constants.CalculatedVariable.OUTLETPRESSURE};R,B=A._pt_analysis_for_vlp(parameters=D,profile_variables=E);B[G]=(B[G]-1.01325)/.980665;B[H]=B[H]*1000000
			if A.gas_well==F:B[I]=B[I]*1000000
		logger.info(f"PT analysis completed");os.makedirs('./pipesim-results',exist_ok=F);S=f"./pipesim-results/{A.well_name}-PT-analysis-report.xlsx"
		if U!='internal-use':
			if U is C:U='Study-1'
			T=U
			if os.path.exists(S):
				while T in list(load_workbook(S).sheetnames):T+='-new'
				with a.ExcelWriter(S,mode='a',engine=d,if_sheet_exists='new')as Y:B.to_excel(Y,sheet_name=T,index=e)
			else:
				with a.ExcelWriter(S,engine=d)as Y:B.to_excel(Y,sheet_name=T,index=e)
			logger.info(f"PT analysis outputs have been written to sheet '{T}' in {S}")
		return B
	except:logger.error('PT analysis failed.')