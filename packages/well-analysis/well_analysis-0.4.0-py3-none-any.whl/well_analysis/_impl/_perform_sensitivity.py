D=round
U=print
T=range
S=int
O=len
N=str
H=float
E=list
B=None
from._logger import logger
from sixgill.definitions import ModelComponents,Parameters,Constants
import pandas as L
from sixgill.pipesim import Model
from sixgill.definitions import Parameters,SystemVariables,ProfileVariables,Constants
import tempfile
from typing import Optional
from sixgill.simulation_result import SimulationResult
import re,pandas as L
from openpyxl import load_workbook
import os
def convert_units(text):
	E=text
	if not isinstance(E,N):return E
	A=E;G=re.search('POUT\\s*=\\s*([\\d.]+)\\s*psia',A)
	if G:B=H(G.group(1));C=S(D(B/14.2233-1.01325));A=re.sub('POUT\\s*=\\s*[\\d.]+\\s*psia',f"POUT={C} ksc",A)
	I=re.search('Flowrate\\s*=\\s*([\\d.]+)\\s*sbbl/day',A)
	if I:B=H(I.group(1));C=B/6.29;A=re.sub('Flowrate\\s*=\\s*[\\d.]+\\s*sbbl/day',f"Flowrate={C:.1f} sm3/d",A)
	J=re.search('INJGAS\\s*=\\s*([\\d.]+)\\s*mmscfd',A)
	if J:B=H(J.group(1));C=S(D(B*28317));A=re.sub('INJGAS\\s*=\\s*[\\d.]+\\s*mmscfd',f"INJGAS={C} SCMD",A)
	F=re.search('(GOR|GLR)\\s*=\\s*([\\d.]+)\\s*scf/sbbl',A)
	if F:K=F.group(1);B=H(F.group(2));C=S(D(B*6.29/35.3147));A=re.sub(f"{K}\\s*=\\s*[\\d.]+\\s*scf/sbbl",f"{K}={C} m3/m3",A)
	return A
def perform_sensitivity(self,study_name=B,thp_sensitivity=B,tubing_sensitivity=B,lift_gas_sensitivity=B,watercut_sensitivity=B,GOR_sensitivity=B,reservoir_pressure_sensitivity=B):
	f='openpyxl';e='main_well';d=GOR_sensitivity;c=watercut_sensitivity;b=True;a='VertComp1';V=study_name;P=self;K=reservoir_pressure_sensitivity;J=lift_gas_sensitivity;I=tubing_sensitivity;G=thp_sensitivity
	try:
		F=[]
		if G is not B:
			G=E(G)
			for C in T(O(G)):G[C]=G[C]*.98066+1.01325
			F.append({Parameters.SystemAnalysisSimulation.SensitivityVariable.COMPONENT:'System Data',Parameters.SystemAnalysisSimulation.SensitivityVariable.VARIABLE:Parameters.SystemAnalysisSimulation.OUTLETPRESSURE,Parameters.SystemAnalysisSimulation.SensitivityVariable.VALUES:G})
		if I is not B:
			I=E(I)
			for C in T(O(I)):I[C]=I[C]*25.4
			F.append({Parameters.SystemAnalysisSimulation.SensitivityVariable.COMPONENT:'Tub1',Parameters.SystemAnalysisSimulation.SensitivityVariable.VARIABLE:Parameters.Tubing.INNERDIAMETER,Parameters.WellPerformanceCurvesSimulation.SensitivityVariable.VALUES:I})
		if J is not B:
			J=E(J)
			for C in T(O(J)):J[C]=J[C]/1000000
			F.append({Parameters.SystemAnalysisSimulation.SensitivityVariable.COMPONENT:'GLV',Parameters.SystemAnalysisSimulation.SensitivityVariable.VARIABLE:Parameters.GasLiftInjection.GASRATE,Parameters.SystemAnalysisSimulation.SensitivityVariable.VALUES:E(J)})
		if c is not B:F.append({Parameters.SystemAnalysisSimulation.SensitivityVariable.COMPONENT:a,Parameters.SystemAnalysisSimulation.SensitivityVariable.VARIABLE:Parameters.Completion.WATERCUT,Parameters.SystemAnalysisSimulation.SensitivityVariable.VALUES:E(c)})
		if d is not B:F.append({Parameters.SystemAnalysisSimulation.SensitivityVariable.COMPONENT:a,Parameters.SystemAnalysisSimulation.SensitivityVariable.VARIABLE:Parameters.Completion.GOR,Parameters.SystemAnalysisSimulation.SensitivityVariable.VALUES:E(d)})
		if K is not B:
			K=E(K)
			for C in T(O(K)):K[C]=K[C]*.98066+1.01325
			F.append({Parameters.SystemAnalysisSimulation.SensitivityVariable.COMPONENT:a,Parameters.SystemAnalysisSimulation.SensitivityVariable.VARIABLE:Parameters.Completion.RESERVOIRPRESSURE,Parameters.SystemAnalysisSimulation.SensitivityVariable.VALUES:K})
		if O(F)==0:logger.info('No sensitivity parameter selected.');return
		g=[SystemVariables.PRESSURE,SystemVariables.TEMPERATURE,SystemVariables.VOLUME_FLOWRATE_LIQUID_STOCKTANK,SystemVariables.VOLUME_FLOWRATE_OIL_STOCKTANK,SystemVariables.VOLUME_FLOWRATE_WATER_STOCKTANK,SystemVariables.VOLUME_FLOWRATE_GAS_STOCKTANK,SystemVariables.GOR_STOCKTANK,SystemVariables.WATER_CUT_STOCKTANK,SystemVariables.VOLUME_FLOWRATE_GAS_STOCKTANK,SystemVariables.WATER_CUT_INSITU,SystemVariables.WELLHEAD_VOLUME_FLOWRATE_FLUID_INSITU,SystemVariables.OUTLET_VOLUME_FLOWRATE_GAS_STOCKTANK,SystemVariables.OUTLET_VOLUME_FLOWRATE_OIL_STOCKTANK,SystemVariables.OUTLET_VOLUME_FLOWRATE_WATER_STOCKTANK,SystemVariables.OUTLET_VOLUME_FLOWRATE_LIQUID_STOCKTANK,SystemVariables.SYSTEM_OUTLET_TEMPERATURE,SystemVariables.BOTTOM_HOLE_PRESSURE,SystemVariables.OUTLET_GLR_STOCKTANK,SystemVariables.OUTLET_WATER_CUT_STOCKTANK];h=[ProfileVariables.TEMPERATURE,ProfileVariables.ELEVATION,ProfileVariables.TOTAL_DISTANCE,ProfileVariables.PRESSURE];i={Parameters.SystemAnalysisSimulation.INLETPRESSURE:P.reservoir_pressure,Parameters.SystemAnalysisSimulation.OUTLETPRESSURE:P.thp,Parameters.SystemAnalysisSimulation.SENSITIVITYMETHOD:Constants.SensitivityMethod.PERMUTED,Parameters.SystemAnalysisSimulation.SENSITIVITYVARIABLES:F};j=P.model.tasks.systemanalysissimulation.run(e,system_variables=g,profile_variables=h,parameters=i);k=P.model.tasks.systemanalysissimulation.validate(e);W=1
		for X in k:U('Issue {}'.format(W));U('Path: {}'.format(X.path));U('Message: {}'.format(X.message));U('Property: {}'.format(X.property_name));W=W+1
		A=L.DataFrame(j.system);A.index=A.index.map(convert_units);l=A.iloc[0];A=A[1:];A.columns=L.MultiIndex.from_arrays([A.columns,l]);m=1000000;n=lambda bara:(bara-1.01325)/.98066
		for Y in A.columns:
			M,D=Y
			if L.isna(D):continue
			if'mmsm3/d'in N(D).lower()or'mmscmd'in N(D).lower():A[M,'SCMD']=A[M,D].astype(H)*m;A.drop(columns=[Y],inplace=b);continue
			if'bara'in N(D).lower():A[M,'ksc']=A[M,D].astype(H).apply(n);A.drop(columns=[Y],inplace=b);continue
			if'scmd'in N(D).lower():A[M,D]=A[M,D].astype(H).round(0).astype(S)
		A.columns=[', '.join(filter(B,A)).strip()for A in A.columns];os.makedirs('./pipesim-results',exist_ok=b);Q=f"./pipesim-results/{P.well_name}-sensitivity-analysis-report.xlsx"
		if V is B:V='Study-1'
		R=V
		if os.path.exists(Q):
			while R in E(load_workbook(Q).sheetnames):R+='-new'
			with L.ExcelWriter(Q,mode='a',engine=f,if_sheet_exists='new')as Z:A.to_excel(Z,sheet_name=R)
		else:
			with L.ExcelWriter(Q,engine=f)as Z:A.to_excel(Z,sheet_name=R)
		logger.info('Sensitivity analysis completed.');logger.info(f"Sensitivity analysis outputs have been written to sheet '{R}' in {Q}")
	except:logger.info('Sensitivity analysis failed!')