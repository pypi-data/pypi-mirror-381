import fnmatch
import os
import re
import subprocess
from collections import deque
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass, field
from enum import Enum, auto
from functools import cached_property
from pathlib import Path
from typing import Any, Optional, Union

from jinja2 import Environment, FileSystemLoader
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from py_app_dev.core.exceptions import UserNotificationException
from py_app_dev.core.subprocess import SubprocessExecutor


class SymbolLinkage(Enum):
    EXTERN = auto()
    LOCAL = auto()


@dataclass
class Symbol:
    name: str
    linkage: SymbolLinkage


@dataclass
class ObjectData:
    path: Path
    symbols: list[Symbol] = field(default_factory=list)

    @property
    def name(self) -> str:
        return self.path.name

    @cached_property
    def required_symbols(self) -> set[str]:
        """Collects all EXTERN symbols into a set."""
        return {symbol.name for symbol in self.symbols if symbol.linkage == SymbolLinkage.EXTERN}

    @cached_property
    def provided_symbols(self) -> set[str]:
        """Collects all LOCAL symbols into a set."""
        return {symbol.name for symbol in self.symbols if symbol.linkage == SymbolLinkage.LOCAL}


class NmExecutor:
    @staticmethod
    def run(obj_file: Path) -> ObjectData:
        obj_data = ObjectData(obj_file)
        executor = SubprocessExecutor(command=["nm", obj_file], capture_output=True, print_output=False)
        completed_process = executor.execute(handle_errors=False)
        if completed_process:
            if completed_process.returncode != 0:
                raise subprocess.CalledProcessError(completed_process.returncode, completed_process.args, stderr=completed_process.stderr)

            # Process the output
            for line in completed_process.stdout.splitlines():
                symbol = NmExecutor.get_symbol(line)
                if symbol:
                    obj_data.symbols.append(symbol)
        else:
            raise UnboundLocalError("nm command failed")
        return obj_data

    @staticmethod
    def get_symbol(nm_symbol_output: str) -> Optional[Symbol]:
        # Regex to capture optional address, mandatory uppercase symbol type, and symbol name
        # Group 1: Symbol Type Letter (e.g., 'U', 'T', 'D', 'B', etc.)
        # Group 2: Symbol Name
        pattern: re.Pattern[str] = re.compile(r"^\s*(?:[0-9A-Fa-f]+\s+)?([A-Z])\s+(\S+)")
        match = pattern.match(nm_symbol_output)

        if match:
            symbol_type_letter = match.group(1)
            symbol_name = match.group(2)

            # Determine linkage based on the symbol type letter
            linkage = SymbolLinkage.EXTERN if symbol_type_letter == "U" else SymbolLinkage.LOCAL

            return Symbol(name=symbol_name, linkage=linkage)

        return None


def parse_objects(obj_files: list[Path], max_workers: Optional[int] = None) -> list[ObjectData]:
    """Run the nm executor on each object file in parallel, collecting all the resulting ObjectData in the same order as `obj_files`."""
    with ThreadPoolExecutor(max_workers=max_workers) as pool:
        # executor.map preserves input order in its output sequence
        results = list(pool.map(NmExecutor.run, obj_files))

    return results


def filter_object_data_symbols(object_data: list[ObjectData], exclude_patterns: list[str] | None) -> list[ObjectData]:
    """
    Filter object symbols by glob patterns returning fresh ``ObjectData`` instances.

    The original ``object_data`` collection is left untouched to avoid stale
    ``cached_property`` values (``required_symbols`` / ``provided_symbols``). By
    reconstructing each ``ObjectData`` we ensure those sets are recomputed from
    the filtered symbol list.

    Args:
        object_data: Parsed object data instances.
        exclude_patterns: Glob (``fnmatch``) patterns for symbols to remove. ``None``
            or an empty list -> no filtering.

    Returns:
        New list with symbols whose names do not match any exclusion pattern.

    """
    if not exclude_patterns:
        return object_data

    filtered: list[ObjectData] = []
    for obj in object_data:
        kept_symbols = [s for s in obj.symbols if not any(fnmatch.fnmatch(s.name, pat) for pat in exclude_patterns)]
        # Recreate ObjectData to avoid stale cached_property values
        filtered.append(ObjectData(path=obj.path, symbols=kept_symbols))
    return filtered


def filter_external_symbols_only(object_data: list[ObjectData]) -> list[ObjectData]:
    """
    Filter object data to keep only symbols with external linkage (EXTERN).

    For dependency analysis, we typically only care about external interfaces
    and dependencies, not internal implementation details (LOCAL symbols).

    Args:
        object_data: Parsed object data instances.

    Returns:
        New list with only symbols that have external linkage (SymbolLinkage.EXTERN).

    """
    filtered: list[ObjectData] = []
    for obj in object_data:
        extern_symbols = [s for s in obj.symbols if s.linkage == SymbolLinkage.EXTERN]
        # Recreate ObjectData to avoid stale cached_property values
        filtered.append(ObjectData(path=obj.path, symbols=extern_symbols))
    return filtered


@dataclass
class DirectoryNode:
    """Represents a node in the directory tree."""

    name: str
    children: dict[str, "DirectoryNode"] = field(default_factory=dict)
    objects: list[ObjectData] = field(default_factory=list)


class DirectoryTreeIterator:
    """
    Iterates over a directory tree structure (dict[str, DirectoryNode]) using BFS.

    Yields a tuple of (parent_name, node), where:
    - parent_name: Optional[str] - The name of the parent node (None for root nodes).
    - node: Union[DirectoryNode, ObjectData] - The current node or object being visited.
    """

    def __init__(self, root_nodes: dict[str, DirectoryNode]):
        self._node_queue: deque[tuple[Optional[str], DirectoryNode]] = deque([(None, node) for node in root_nodes.values()])  # Queue for nodes to visit, with parent name
        self._object_queue: deque[tuple[str, ObjectData]] = deque()  # Queue for objects to yield from the current node

    def __iter__(self) -> "DirectoryTreeIterator":
        return self

    def __next__(self) -> tuple[Optional[str], Union[DirectoryNode, ObjectData]]:
        while True:
            # If there are objects buffered from the last node, yield them first
            if self._object_queue:
                return self._object_queue.popleft()

            # If node queue is empty, iteration is done
            if not self._node_queue:
                raise StopIteration

            # Process the next node in the queue
            parent_name, current_node = self._node_queue.popleft()

            # Add all objects from this node to the object queue
            for obj in current_node.objects:
                self._object_queue.append((current_node.name, obj))

            # Add all children nodes to the node queue for future processing
            for _, child_node in current_node.children.items():
                self._node_queue.append((current_node.name, child_node))

            # Yield the current node itself
            return (parent_name, current_node)


class ObjectsDependenciesReportGenerator:
    def __init__(self, object_data: list[ObjectData], use_parent_deps: bool = False):
        self.object_data = object_data
        self.use_parent_deps = use_parent_deps

    def generate_report(self, output_file: Path) -> None:
        """Generates the HTML report by rendering the Jinja2 template with the graph data."""
        graph_data = self.generate_graph_data()

        env = Environment(loader=FileSystemLoader(Path(__file__).parent), autoescape=True)
        template = env.get_template("object_analyzer.html.jinja")
        rendered_html = template.render(graph_data=graph_data)

        # Write the rendered HTML to the output file
        with open(output_file, "w", encoding="utf-8") as file:
            file.write(rendered_html)

    @staticmethod
    def _build_directory_tree(objects: list[ObjectData]) -> dict[str, DirectoryNode]:
        """Builds a nested structure representing the directory hierarchy, stopping at 'CMakeFiles'."""
        common_path = ObjectsDependenciesReportGenerator._determine_common_path([obj.path.absolute() for obj in objects])
        root_nodes: dict[str, DirectoryNode] = {}

        for obj in objects:
            obj_relative_path = obj.path.parent.resolve().relative_to(common_path)
            parts = list(obj_relative_path.parts)
            parts = parts[: parts.index("CMakeFiles")]  # Keep parts up to, but not including, CMakeFiles

            # Build tree structure and populate objects
            current_children = root_nodes
            current_node: Optional[DirectoryNode] = None

            for part in parts:
                if part not in current_children:
                    new_node = DirectoryNode(name=part)
                    current_children[part] = new_node
                current_node = current_children[part]
                current_children = current_node.children  # Move deeper
            # Add the object to the node corresponding to the final directory part
            if current_node:
                current_node.objects.append(obj)

        return root_nodes

    @staticmethod
    def _determine_common_path(objects_paths: list[Path]) -> Path:
        # Find Common Base Path
        try:
            common_base_str = os.path.commonpath([str(path) for path in objects_paths])
            common_base = Path(common_base_str)
        except ValueError:
            raise UserNotificationException("No common base path found. Ensure all object files are in the same directory tree.") from None

        # Check if the common base already contains "CMakeFiles" and return the parent parent directory of "CMakeFiles"
        if "CMakeFiles" in common_base.parts:
            cmakefiles_index = common_base.parts.index("CMakeFiles")
            common_base = Path(*common_base.parts[:cmakefiles_index]).parent
        else:
            # We need to make sure that there will be at least one root node.
            # If all objects are in the same directory the common path will be up to the `CMakeFiles` directory and we will end up with no root node!
            found_empty_path = False
            for obj_path in objects_paths:
                obj_relative_path = obj_path.parent.resolve().relative_to(common_base)
                parts = list(obj_relative_path.parts)
                # Check if the path contains "CMakeFiles"
                try:
                    cmakefiles_index = parts.index("CMakeFiles")
                    parts = parts[:cmakefiles_index]  # Keep parts up to, but not including, CMakeFiles
                except ValueError:
                    raise UserNotificationException(f"'CMakeFiles' directory not found in the path {obj_relative_path}.") from None
                if not parts:
                    found_empty_path = True
                    break
            if found_empty_path:
                common_base = common_base.parent
        return common_base

    def generate_graph_data(self) -> dict[str, list[dict[str, Any]]]:
        """Converts a list of ObjectData into a dictionary suitable for Cytoscape.js containing nodes and edges representing object dependencies."""
        objects = self.object_data
        edges = []
        edge_set = set()  # To avoid duplicate edges between the same pair
        node_connections = {obj.name: 0 for obj in objects}

        # Determine edges and count connections
        for i in range(len(objects)):
            for j in range(i + 1, len(objects)):
                obj1 = objects[i]
                obj2 = objects[j]

                # Check if obj1 provides any symbol required by obj2
                provides_required = bool(obj1.provided_symbols.intersection(obj2.required_symbols))
                # Check if obj2 provides any symbol required by obj1
                required_provides = bool(obj2.provided_symbols.intersection(obj1.required_symbols))

                if provides_required or required_provides:
                    # Ensure edge ID is consistent regardless of order
                    source_name, target_name = sorted([obj1.name, obj2.name])
                    edge_id = f"{source_name}.{target_name}"

                    if edge_id not in edge_set:
                        edges.append(
                            {
                                "data": {
                                    "id": edge_id,
                                    "source": obj1.name,  # Cytoscape doesn't strictly enforce source/target for undirected
                                    "target": obj2.name,
                                }
                            }
                        )
                        edge_set.add(edge_id)
                        node_connections[obj1.name] += 1
                        node_connections[obj2.name] += 1
        nodes = []
        root_nodes = ObjectsDependenciesReportGenerator._build_directory_tree(objects)
        # Iterate over the whole node tree to get the nodes
        for parent_name, node in DirectoryTreeIterator(root_nodes):
            if isinstance(node, ObjectData):
                obj = node
                node_size = 5 + node_connections[obj.name] * 2  # Example scaling
                data = {
                    "data": {
                        "id": obj.name,
                        "size": node_size,
                        "content": obj.name,  # Display object name
                        "font_size": 10,  # Adjust font size as needed
                    }
                }
                if parent_name and self.use_parent_deps:
                    data["data"]["parent"] = parent_name
                nodes.append(data)
            # This is a DirectoryNode
            else:
                data = {
                    "data": {
                        "id": node.name,
                        "size": 5,  # Base size for root nodes
                        "content": node.name,  # Display object name
                        "font_size": 10,  # Adjust font size as needed
                        # TODO: populate parent
                    }
                }
                if parent_name and self.use_parent_deps:
                    data["data"]["parent"] = parent_name
                nodes.append(data)
        return {"nodes": nodes, "edges": edges}


class ExcelColumnMapper:
    """Manages Excel column mappings for the Objects sheet."""

    def __init__(self) -> None:
        # Define column positions
        self.OBJECT_NAME = 1
        self.FILE_PATH = 2
        self.PROVIDED_TOTAL = 3
        self.PROVIDED_USED = 4
        self.PROVIDED_DEPENDENCIES = 5
        self.REQUIRED_COUNT = 6
        self.REQUIRED_DEPENDENCIES = 7
        self.TOTAL_SYMBOLS = 8

        # Define column groups for headers
        self.OBJECT_INFO_START = self.OBJECT_NAME
        self.OBJECT_INFO_END = self.FILE_PATH

        self.PROVIDED_INTERFACES_START = self.PROVIDED_TOTAL
        self.PROVIDED_INTERFACES_END = self.PROVIDED_DEPENDENCIES

        self.REQUIRED_INTERFACES_START = self.REQUIRED_COUNT
        self.REQUIRED_INTERFACES_END = self.REQUIRED_DEPENDENCIES

        self.TOTAL_COLUMN = self.TOTAL_SYMBOLS

        # Total number of columns
        self.TOTAL_COLUMNS = self.TOTAL_SYMBOLS

        # Header texts
        self.MAIN_HEADERS = {
            self.OBJECT_INFO_START: "Object Info",
            self.PROVIDED_INTERFACES_START: "Provided Interfaces",
            self.REQUIRED_INTERFACES_START: "Required Interfaces",
            self.TOTAL_COLUMN: "Total",
        }

        self.SUB_HEADERS = [
            "Object Name",  # OBJECT_NAME
            "File Path",  # FILE_PATH
            "Total",  # PROVIDED_TOTAL
            "Used",  # PROVIDED_USED
            "Dependencies",  # PROVIDED_DEPENDENCIES
            "Count",  # REQUIRED_COUNT
            "Dependencies",  # REQUIRED_DEPENDENCIES
            "Symbols",  # TOTAL_SYMBOLS
        ]

        # Column ranges for merging
        self.MERGE_RANGES = [
            "A1:B1",  # Object Info
            "C1:E1",  # Provided Interfaces
            "F1:G1",  # Required Interfaces
        ]

        # Columns that get main headers (for styling)
        self.MAIN_HEADER_COLUMNS = [self.OBJECT_INFO_START, self.PROVIDED_INTERFACES_START, self.REQUIRED_INTERFACES_START, self.TOTAL_COLUMN]


class ObjectsDataExcelReportGenerator:
    """Excel report generator for object data with dependency analysis."""

    def __init__(
        self,
        object_data: list[ObjectData],
        use_parent_deps: bool = False,
        create_traceability_matrix: bool = False,
    ) -> None:
        self.object_data = object_data
        self.use_parent_deps = use_parent_deps
        self.create_traceability_matrix = create_traceability_matrix
        self.columns = ExcelColumnMapper()

    def generate_report(self, output_file: Path) -> None:
        wb = Workbook()
        self._create_objects_sheet(wb)
        if self.create_traceability_matrix:
            self._create_dependency_matrix_sheet(wb)
        wb.save(output_file)

    def _create_objects_sheet(self, wb: Workbook) -> None:
        """Create the Objects sheet with object statistics."""
        ws = wb.active
        if ws is None:
            raise RuntimeError("Failed to create worksheet")
        ws.title = "Objects"

        # Create grouped headers
        self._create_grouped_headers(ws, self.columns)

        for row, obj in enumerate(self.object_data, 3):  # Start from row 3 due to grouped headers
            # Calculate dependencies
            objects_requiring_from_this = [other_obj.name for other_obj in self.object_data if other_obj != obj and obj.provided_symbols.intersection(other_obj.required_symbols)]
            objects_providing_to_this = [other_obj.name for other_obj in self.object_data if other_obj != obj and other_obj.provided_symbols.intersection(obj.required_symbols)]

            # Calculate used provided interfaces (symbols from this object that are actually required by other objects)
            used_provided_interfaces = set()
            for other_obj in self.object_data:
                if other_obj != obj:
                    used_provided_interfaces.update(obj.provided_symbols.intersection(other_obj.required_symbols))

            ws.cell(row=row, column=self.columns.OBJECT_NAME, value=obj.name)
            ws.cell(row=row, column=self.columns.FILE_PATH, value=str(obj.path))
            ws.cell(row=row, column=self.columns.PROVIDED_TOTAL, value=len(obj.provided_symbols))
            ws.cell(row=row, column=self.columns.PROVIDED_USED, value=len(used_provided_interfaces))
            ws.cell(row=row, column=self.columns.PROVIDED_DEPENDENCIES, value="\n".join(objects_requiring_from_this) if objects_requiring_from_this else "None")
            ws.cell(row=row, column=self.columns.REQUIRED_COUNT, value=len(obj.required_symbols))
            ws.cell(row=row, column=self.columns.REQUIRED_DEPENDENCIES, value="\n".join(objects_providing_to_this) if objects_providing_to_this else "None")
            ws.cell(row=row, column=self.columns.TOTAL_SYMBOLS, value=len(obj.symbols))

        self._auto_adjust_columns(ws, self.columns.TOTAL_COLUMNS)

    def _create_dependency_matrix_sheet(self, wb: Workbook) -> None:
        """Create the Dependency Matrix sheet showing interface dependencies."""
        ws = wb.create_sheet("Dependency Matrix")

        interface_usage = self._calculate_interface_usage()
        obj_names = [obj.name for obj in self.object_data]

        fixed_headers = ["Object", "Interface", "Usage Count"]
        headers = [*fixed_headers, *obj_names]
        self._create_header_row(ws, headers)

        # Freeze the first row and fixed headers
        freeze_column = get_column_letter(len(fixed_headers) + 1)
        ws.freeze_panes = f"{freeze_column}2"

        current_row = 2
        for obj in self.object_data:
            if not obj.provided_symbols:
                continue

            for interface_name in sorted(obj.provided_symbols):
                self._create_dependency_row(ws, current_row, obj.name, interface_name, interface_usage[interface_name])
                current_row += 1

        self._auto_adjust_columns(ws, len(headers))

    def _create_header_row(self, ws: Any, headers: list[str]) -> None:
        """Create and style header row."""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _create_dependency_row(self, ws: Any, row: int, obj_name: str, interface_name: str, usage_count: int) -> None:
        """Create a single dependency matrix row."""
        ws.cell(row=row, column=1, value=obj_name).font = Font(bold=True)
        ws.cell(row=row, column=2, value=interface_name)

        usage_cell = ws.cell(row=row, column=3, value=usage_count)
        usage_cell.font = Font(bold=usage_count > 0)
        usage_cell.alignment = Alignment(horizontal="center", vertical="center")

        for col, requiring_obj in enumerate(self.object_data, 4):
            if interface_name in requiring_obj.required_symbols:
                cell = ws.cell(row=row, column=col, value="X")
                cell.font = Font(bold=True, color="FF0000")
                cell.alignment = Alignment(horizontal="center", vertical="center")

    def _calculate_interface_usage(self) -> dict[str, int]:
        """Calculate usage count for each interface."""
        interface_usage = {}

        for obj in self.object_data:
            for symbol in obj.provided_symbols:
                usage_count = sum(1 for other_obj in self.object_data if symbol in other_obj.required_symbols)
                interface_usage[symbol] = usage_count

        return interface_usage

    def _auto_adjust_columns(self, ws: Any, column_count: int) -> None:
        """Auto-adjust column widths."""
        for col in range(1, column_count + 1):
            ws.column_dimensions[get_column_letter(col)].auto_size = True

    def _create_grouped_headers(self, ws: Any, column_mapper: ExcelColumnMapper) -> None:
        """Create grouped headers with main categories and sub-headers."""
        # First row - main category headers
        for start_col in column_mapper.MAIN_HEADER_COLUMNS:
            ws.cell(row=1, column=start_col, value=column_mapper.MAIN_HEADERS[start_col])

        # Merge cells for main headers
        for merge_range in column_mapper.MERGE_RANGES:
            ws.merge_cells(merge_range)

        # Second row - specific column headers
        sub_headers = column_mapper.SUB_HEADERS

        # Style main headers
        for col in column_mapper.MAIN_HEADER_COLUMNS:
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill(start_color="B0B0B0", end_color="B0B0B0", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Style and populate sub-headers
        for col, header in enumerate(sub_headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add borders to separate groups
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row in range(1, 3):
            for col in range(1, column_mapper.TOTAL_COLUMNS + 1):
                ws.cell(row=row, column=col).border = thin_border
