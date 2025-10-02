from __future__ import annotations
from typing import Iterable
from ..types import UnifiedDocument, Metadata, Section, Chunk
from ..core import register_parser, ParserProtocol

class ExcelParser(ParserProtocol):
    name = "excel"
    content_types: Iterable[str] = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel"
    )
    extensions: Iterable[str] = (".xlsx", ".xls")

    def can_parse(self, meta: Metadata) -> bool:
        return (meta.content_type in self.content_types or 
                (meta.path or "").endswith((".xlsx", ".xls")))

    def parse(self, target, meta: Metadata, recursive: bool = False, **kwargs) -> UnifiedDocument:
        doc = UnifiedDocument(meta=meta, sections=[])
        
        try:
            import openpyxl
            from openpyxl import load_workbook
            
            # Load workbook
            workbook = load_workbook(str(target), data_only=True)
            
            # Extract workbook properties
            if workbook.properties.title:
                doc.meta.title = workbook.properties.title
            if workbook.properties.creator:
                doc.meta.extra["creator"] = workbook.properties.creator
            if workbook.properties.created:
                doc.meta.extra["created"] = workbook.properties.created.isoformat()
            if workbook.properties.modified:
                doc.meta.extra["modified"] = workbook.properties.modified.isoformat()
            
            order = 0
            
            # Process each worksheet
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # Extract sheet data
                sheet_data = []
                max_row = worksheet.max_row
                max_col = worksheet.max_column
                
                for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True):
                    # Convert None values to empty strings and filter out completely empty rows
                    row_data = [str(cell) if cell is not None else "" for cell in row]
                    if any(cell.strip() for cell in row_data):  # Only add non-empty rows
                        sheet_data.append(" | ".join(row_data))
                
                if sheet_data:
                    sheet_text = "\n".join(sheet_data)
                    doc.sections.append(Section(
                        heading=f"Sheet: {sheet_name}",
                        chunks=[Chunk(text=sheet_text, order=order, meta={"type": "sheet"})],
                        meta={"type": "sheet", "sheet_name": sheet_name, "rows": len(sheet_data), "cols": max_col}
                    ))
                    order += 1
            
            # Store workbook info
            doc.meta.extra["workbook_info"] = {
                "sheet_count": len(workbook.sheetnames),
                "sheet_names": workbook.sheetnames
            }
            
        except ImportError:
            error_text = "[panparsex:excel] openpyxl not available. Install with: pip install openpyxl"
            doc.sections.append(Section(
                heading="Error",
                chunks=[Chunk(text=error_text, order=0, meta={"error": True})],
                meta={"error": True}
            ))
        except Exception as e:
            error_text = f"[panparsex:excel] unable to parse: {e}"
            doc.sections.append(Section(
                heading="Error",
                chunks=[Chunk(text=error_text, order=0, meta={"error": True})],
                meta={"error": True}
            ))
        
        return doc

register_parser(ExcelParser())
